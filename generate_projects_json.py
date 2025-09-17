import os, yaml, json
import pandas as pd
from collections import OrderedDict
from openpyxl import load_workbook

project_dir = "projects"
output_file = "dashboard/projects.json"
all_metadata = []

created_count = 0
missing_projects = []

# read .gitignore and build pathspec to ignore when compiling projects
import pathspec
with open('.gitignore') as f:
    gitignore = f.read()
spec = pathspec.PathSpec.from_lines('gitwildmatch', gitignore.splitlines())

# iterate project folders to build metadata
for folder in os.listdir(project_dir):
    if folder.startswith("."):
        continue
    folder_path = os.path.join(project_dir, folder)
    if not os.path.isdir(folder_path):
        continue
    metadata_path = os.path.join(folder_path, "metadata.yml")

    # is metadata yaml missing? convert from .xlsx or .csv
    # auto-convert any metadata .xlsx or .csv (filename must contain "metadata") -> metadata.yml
    src = next((f for f in os.listdir(folder_path)
                if not f.startswith("~$") and "metadata" in f.lower() and os.path.splitext(f)[1].lower() in (".xlsx", ".csv")), None)
    if src:
        src_path = os.path.join(folder_path, src)
        ext = os.path.splitext(src)[1].lower()
        meta_dict = {}
        built_direct = False
        # read key/value rows where col A = key, cols B.. = values (no header)
        if ext == ".xlsx":
            # Build meta_dict directly from openpyxl rows (A=key, B..=values)
            wb = load_workbook(src_path, read_only=True, data_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                if r is None:
                    continue
                cells = list(r)
                if not cells or (cells[0] is None or str(cells[0]).strip() == ""):
                    continue
                key = str(cells[0]).strip()
                vals = [v for v in cells[1:] if v is not None and str(v).strip() != ""]
                if not vals:
                    continue
                if len(vals) == 1:
                    s = str(vals[0]).strip()
                    if s.startswith("[") and s.endswith("]"):
                        try:
                            meta_dict[key] = yaml.safe_load(s)
                            continue
                        except Exception:
                            pass
                    meta_dict[key] = s
                else:
                    meta_dict[key] = [str(v).strip() for v in vals]
            built_direct = True
        else:
            df = pd.read_csv(src_path, header=None)

        # if not built directly, use pandas DataFrame path
        if not built_direct:
            meta_dict = {}
            for _, row in df.iterrows():
                if row.isna().all():
                    continue
                key_cell = row.iloc[0]
                key = str(key_cell).strip() if pd.notna(key_cell) else ""
                if not key or key.lower() in ("nan", "none"):
                    continue
                vals = [v for v in row.iloc[1:].tolist() if pd.notna(v) and str(v).strip() != ""]
                if not vals:
                    continue
                if len(vals) == 1:
                    s = str(vals[0]).strip()
                    if s.startswith("[") and s.endswith("]"):
                        try:
                            meta_dict[key] = yaml.safe_load(s)
                            continue
                        except Exception:
                            pass
                    meta_dict[key] = s
                else:
                    meta_dict[key] = [str(v).strip() for v in vals]
        # normalize case for specific fields: data, methods, themes
        def _norm_case(val):
            if isinstance(val, str):
                return val if val.isupper() or val.islower() else val.lower()
            if isinstance(val, list):
                out = []
                for x in val:
                    if isinstance(x, str):
                        out.append(x if x.isupper() or x.islower() else x.lower())
                    elif x is not None:
                        out.append(str(x))
                return out
            return val
        for _fld in ("data", "methods", "themes"):
            if _fld in meta_dict:
                meta_dict[_fld] = _norm_case(meta_dict[_fld])

        # order keys for output
        _order = ["title", "authors", "language", "data", "methods", "themes", "description"]
        ordered = OrderedDict()
        for k in _order:
            if k in meta_dict:
                ordered[k] = meta_dict[k]
        for k, v in meta_dict.items():
            if k not in ordered:
                ordered[k] = v

        # ensure YAML flow-list for list-like fields (e.g., ["a", "b"]) and coerce to list
        class _FlowList(list):
            pass

        def _flow_list_representer(dumper, seq):
            return dumper.represent_sequence('tag:yaml.org,2002:seq', seq, flow_style=True)

        yaml.SafeDumper.add_representer(_FlowList, _flow_list_representer)

        class _Quoted(str):
            pass

        def _quoted_representer(dumper, data):
            return dumper.represent_scalar('tag:yaml.org,2002:str', str(data), style='"')

        yaml.SafeDumper.add_representer(_Quoted, _quoted_representer)

        ordered_plain = dict(ordered)

        def _to_quoted_list(v):
            if isinstance(v, str):
                parts = [p.strip() for p in v.split(",")]
                vals = parts if len(parts) > 1 else [v]
            elif isinstance(v, list):
                vals = [str(x) for x in v]
            else:
                vals = [str(v)]
            return _FlowList([_Quoted(x) for x in vals])

        for k, v in list(ordered_plain.items()):
            if v is None:
                continue
            if k in ("themes", "methods"):
                ordered_plain[k] = _to_quoted_list(v)
            else:
                if isinstance(v, list):
                    ordered_plain[k] = [_Quoted(str(x)) for x in v]
                else:
                    ordered_plain[k] = _Quoted(str(v))

        with open(metadata_path, "w") as f:
            yaml.safe_dump(ordered_plain, f, allow_unicode=True, sort_keys=False)

        created_count += 1

    # if metadata.yml exists, read and append to all_metadata
    if os.path.isfile(metadata_path):
        with open(metadata_path, 'r') as f:
            metadata = yaml.safe_load(f)
        
        metadata["folder"] = folder
        # add relative file paths for all files except metadata.yml / metadata.xlsx
        files = [
            f for f in os.listdir(folder_path)
            if f not in ("metadata.yml", src)
            # drop gitignored files
            and not spec.match_file(os.path.join(folder_path, f))
        ]
        metadata["files"] = [f"{folder}/{f}" for f in files]
        all_metadata.append(metadata)
    else:
        missing_projects.append(folder)

with open(output_file, 'w') as f:
    json.dump(all_metadata, f, indent=2)

print(f"{created_count} metadata yml files created.")
if missing_projects:
    print("Projects missing metadata:", ", ".join(missing_projects))